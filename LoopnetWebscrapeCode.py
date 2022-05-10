# -*- coding: utf-8 -*-
"""
Created on Thu Mar 17 15:09:21 2022

Entire Loopnet Scrape Code

@author: Kyle
"""

#part 1: Use input from excel to search for subject property

import openpyxl
from openpyxl import Workbook
wb = Workbook()
wb = openpyxl.load_workbook(filename = "AutomaticPropertyUnderwriter.xlsm", read_only=False, keep_vba=True)
ws1 = wb.worksheets[0]
Input_value = ws1["A1"].value
num_comps = ws1["A2"].value

#part2: webscrape details of the subject property

import requests
from bs4 import BeautifulSoup
import pandas as pd
from time import sleep
from numpy import random
import re

headers = {'User-Agent': 'Mozilla/5.0 (Windows NT 6.1; WOW64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/56.0.2924.76 Safari/537.36', "Upgrade-Insecure-Requests": "1","DNT": "1","Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8","Accept-Language": "en-US,en;q=0.5","Accept-Encoding": "gzip, deflate"}
url = str(Input_value)

page = requests.get(url, headers=headers)

soup = BeautifulSoup(page.text, 'lxml')

title = soup.find('h1', class_='profile-hero-title').text.strip()
link = soup.find('a', class_= 'button text print ldp-header-nav-print').get('href')
link = 'https://www.loopnet.com/' + link[:-5]
state = soup.find_all('a', class_='breadcrumbs__crumb-link')[1].text.strip()
town = soup.find_all('a', class_='breadcrumbs__crumb-link')[2].text.strip()

wb = openpyxl.load_workbook(filename = "OutputData.xlsx", read_only=False)
ws1 = wb.worksheets[4]
ws1['A1'] = title
ws1['A2'] = link
ws1['A3'] = state
ws1['A4'] = town
wb.save("OutputData.xlsx")

table_key = 0

table = soup.find('div', class_='property-facts__facts-wrap property-data')
try:
    first_row = table.find('div', class_= 'property-facts-row')
except:
    table = soup.find('table', class_='property-data featured-grid')
    table_key = 1
    
if table_key != 1:
    try:
        Asset_Class = table.find(string = re.compile('Property Type')).find_next('div').text.strip()
    except:
        Asset_Class = 'null'
else:
    try:
        Asset_Class = table.find(string = re.compile('Property Type')).find_next('td').text.strip()
    except:
        Asset_Class = 'null'
    
try:
    feature_section = soup.find('div', class_='property-facts__facts-wrap property-data')
    df1 = pd.DataFrame({'Property fact':[''], 'Data':['']})
    facts = feature_section.find_all('div', class_= 'property-fact-value-container')
    for fact in facts:
        property_fact = fact.find('div', class_= 'fact-name').text.strip()
        property_data = fact.find('div', class_= 'property-facts-value-items').text.strip()
        df1 = df1.append({'Property fact':property_fact, 'Data':property_data}, ignore_index = True)
except:
    try:
        feature_section = soup.find('table', class_='property-data featured-grid')
        df1 = pd.DataFrame({'Property fact':[''], 'Data':['']})
        facts = feature_section.find_all('tr')
        for fact in facts:
            property_fact = fact.find_all('td')[0].text.strip()
            property_data = fact.find_all('td')[1].text.strip()
            df1 = df1.append({'Property fact':property_fact, 'Data':property_data}, ignore_index = True)
            property_fact = fact.find_all('td')[2].text.strip()
            property_data = fact.find_all('td')[3].text.strip()
            df1 = df1.append({'Property fact':property_fact, 'Data':property_data}, ignore_index = True)
    except:
        details = {
        'Property Fact' : ['NA'],
        'Data' : ['NA']
        }
        df1 = pd.DataFrame(details)

try:    
    table = soup.find('table', class_='property-data summary')
    headers = []
    column_headers = table.find_all('th')
    for i in column_headers:
        header = i.text.strip()
        headers.append(header)
    df2 = pd.DataFrame(columns = headers)
    for j in table.find_all('tr')[1:]:
        row_data = j.find_all('td')
        row = [tr.text.strip() for tr in row_data]
        length = len(df2)
        df2.loc[length] = row
except:
    details = {
    'Property Fact' : ['NA'],
    'Data' : ['NA']
    }
    df2 = pd.DataFrame(details)

try:
    table = soup.find('table', class_='property-data summary financial')
    headers = []
    column_headers = table.find_all('th')
    for i in column_headers:
        header = i.text.strip()
        headers.append(header)
    df3 = pd.DataFrame(columns = headers)
    for j in table.find_all('tr')[1:]:
        row_data = j.find_all('td')
        row = [tr.text.strip() for tr in row_data]
        length = len(df3)
        df3.loc[length] = row
except:
    details = {
    'Property Fact' : ['NA'],
    'Data' : ['NA']
    }
    df3 = pd.DataFrame(details)

#part 3: Get to the correct page to find a sample of properties near subject property

us_state_to_abbrev = {
    "Alabama": "AL",
    "Alaska": "AK",
    "Arizona": "AZ",
    "Arkansas": "AR",
    "California": "CA",
    "Colorado": "CO",
    "Connecticut": "CT",
    "Delaware": "DE",
    "Florida": "FL",
    "Georgia": "GA",
    "Hawaii": "HI",
    "Idaho": "ID",
    "Illinois": "IL",
    "Indiana": "IN",
    "Iowa": "IA",
    "Kansas": "KS",
    "Kentucky": "KY",
    "Louisiana": "LA",
    "Maine": "ME",
    "Maryland": "MD",
    "Massachusetts": "MA",
    "Michigan": "MI",
    "Minnesota": "MN",
    "Mississippi": "MS",
    "Missouri": "MO",
    "Montana": "MT",
    "Nebraska": "NE",
    "Nevada": "NV",
    "New Hampshire": "NH",
    "New Jersey": "NJ",
    "New Mexico": "NM",
    "New York": "NY",
    "North Carolina": "NC",
    "North Dakota": "ND",
    "Ohio": "OH",
    "Oklahoma": "OK",
    "Oregon": "OR",
    "Pennsylvania": "PA",
    "Rhode Island": "RI",
    "South Carolina": "SC",
    "South Dakota": "SD",
    "Tennessee": "TN",
    "Texas": "TX",
    "Utah": "UT",
    "Vermont": "VT",
    "Virginia": "VA",
    "Washington": "WA",
    "West Virginia": "WV",
    "Wisconsin": "WI",
    "Wyoming": "WY",
    "District of Columbia": "DC",
    "American Samoa": "AS",
    "Guam": "GU",
    "Northern Mariana Islands": "MP",
    "Puerto Rico": "PR",
    "United States Minor Outlying Islands": "UM",
    "U.S. Virgin Islands": "VI",
}

Asset_Class_Search_Keys = {'Office':'office-buildings', 'Industrial':'industrial-properties', 'Retail':'retail-properties', 'Restaurant':'restaurants',
                           'Shopping Center':'shopping-centers', 'Multifamily':'apartment-buildings', 'Specialty':'specialty-properties', 'Health Care': 'health-care-facilities',
                           'Hospitality':'hospitality-properties', 'Sports & Entertainment':'sports-entertainment-properties', 'Land':'land', 'Residential Income':'residential-income-properties',
                           'null':'commercial-real-estate'}

try:
    url = 'https://www.loopnet.com/search/' + Asset_Class_Search_Keys[Asset_Class] + '/' + str(town)+ '-' + str(us_state_to_abbrev[state]) + '/for-sale/?'
except:
    url = 'https://www.loopnet.com/search/' + Asset_Class_Search_Keys['null'] + '/' + str(town)+ '-' + str(us_state_to_abbrev[state]) + '/for-sale/?'
from selenium import webdriver

driver = webdriver.Chrome(r'C:\Users\Kyle\Downloads\chromedriver.exe')
driver.get(url)
sleep(random.uniform(1, 2))

soup = BeautifulSoup(driver.page_source, 'lxml')

try:
    listing_count = int(driver.find_element_by_xpath('//*[@id="placardSec"]/div[3]/div/ol/div/span').text.split(' ')[-1])
except:
    listing_count = 0

if listing_count < num_comps:
    rb_button = driver.find_element_by_xpath('//*[@id="mapSec"]/div[1]/div[2]')
    rb_button.click()
    sleep(random.uniform(3, 4))

soup = BeautifulSoup(driver.page_source, 'lxml')    

try:
    listing_count = int(driver.find_element_by_xpath('//*[@id="placardSec"]/div[3]/div/ol/div/span').text.split(' ')[-1])
except:
    listing_count = 0
    
while listing_count < num_comps:
    zo_button = driver.find_element_by_xpath('//*[@id="mapSec"]/div[1]/div[1]/ul/li[2]')
    zo_button.click()
    sleep(random.uniform(3, 4))
    soup = BeautifulSoup(driver.page_source, 'lxml')
    try:
        listing_count = int(driver.find_element_by_xpath('//*[@id="placardSec"]/div[3]/div/ol/div/span').text.split(' ')[-1])
    except:
        listing_count = 0
#part 4: Loop through all the properties and make a dataframe of all the details

soup = BeautifulSoup(driver.page_source, 'lxml')
first_link = soup.find('div', class_='placard-pseudo')
first_property = first_link.find('a').get('href')
driver.get(first_property)
sleep(random.uniform(3, 4))
soup = BeautifulSoup(driver.page_source, 'lxml')

df4 = pd.DataFrame({'address':[''], 'link':[''], 'size':[''], 'price':[''], 'cap rate':[''], 'number of units':[''], 'year built':[''], 'class':[''],
                   'GRI':[''], 'vacancy':[''], 'taxes':[''], 'OPEX':[''], 'NOI':[''], 'zip code':['']})
pd.options.mode.use_inf_as_na = True

for n in range(num_comps):
    try:
        address = soup.find('h1', class_='profile-hero-title').text.strip()
    except:
        address = soup.find('div', class_='column-08 property-info-title').find_next('h1').text.strip()
        
    link = soup.find('a', class_= 'button text print ldp-header-nav-print').get('href')
    link = 'https://www.loopnet.com/' + link[:-5]
    
    try:    
        zipo = soup.find('div', class_= 'breadcrumbs__crumbs')
        zips = zipo.find('h1', class_= 'breadcrumbs__crumb breadcrumbs__crumb-title')
        zip_code = zips.find('a', class_= 'breadcrumbs__crumb-link').text.strip()
    except:
        zip_code = 'NA'
    
    table_key = 0
    
    table = soup.find('div', class_='property-facts__facts-wrap property-data')
    try:
        first_row = table.find('div', class_= 'property-facts-row')
    except:
        table = soup.find('table', class_='property-data featured-grid')
        table_key = 1
        
    if table_key != 1:
        try: 
            size1 = table.find(string = re.compile('Building Size')).find_next('div').text.strip() 
            size = ''
            for c in size1:
                if c.isdigit():
                    size = size + c
        except:
            size = 'NA'
        try:
            price1 = table.find(string = re.compile('Price')).find_next('div').text.strip() 
            price = ''
            for c in price1:
                if c.isdigit():
                    price = price + c
        except:
            price = 'NA'
        try:
            cap_rate = table.find(string = re.compile('Cap')).find_next('div').text.strip() 
        except:
            cap_rate = 'NA'
        try:
            num_units = table.find(string = re.compile('No. Units')).find_next('div').text.strip()
        except:
            num_units = 'NA'
        try:
            year_built = table.find(string = re.compile('Year Built')).find_next('div').text.strip()
            if "/" in year_built:
                year_built = year_built.split("/")
                year_built = int(year_built[0])
        except:
            year_built = 'NA'
        try:
            Class = table.find(string = re.compile('Building Class')).find_next('div').text.strip()
        except:
            Class = 'NA'
        
        if Class != 'NA':
            if Class == "A":
                Class = 3
            elif Class == "B":
                Class = 2
            elif Class == "C":
                Class = 1
            else:
                Class = "NA"
        print('went opetion 1')
    else:
        try: 
            size1 = table.find(string = re.compile('Building Size')).find_next('td').text.strip()
            size = ''
            for c in size1:
                if c.isdigit():
                    size = size + c
        except:
            size = 'NA'
        try:
            price1 = table.find(string = re.compile('Price')).find_next('td').text.strip()
            price = ''
            for c in price1:
                if c.isdigit():
                    price = price + c
        except:
            price = 'NA'
        try:
            cap_rate = table.find(string = re.compile('Cap')).find_next('td').text.strip() 
        except:
            cap_rate = 'NA'
        try:
            num_units = table.find(string = re.compile('No. Units')).find_next('td').text.strip()
        except:
            num_units = 'NA'
        try:
            year_built = table.find(string = re.compile('Year Built')).find_next('td').text.strip()
            if "/" in year_built:
                year_built = year_built.split("/")
                year_built = int(year_built[0])
        except:
            year_built = 'NA'
        try:
            Class = table.find(string = re.compile('Building Class')).find_next('td').text.strip()
        except:
            Class = 'NA'
        
        if Class != 'NA':
            if Class == "A":
                Class = 3
            elif Class == "B":
                Class = 2
            elif Class == "C":
                Class = 1
            else:
                Class = "NA"
        
        print('went option 2')
#end paste


#New Table

    try:
        table = soup.find('table', class_= 'property-data summary financial')
        
        GRI = table.find(string = re.compile('Gross Rental Income')).find_next('td').text.strip()
    except:
        GRI = 'NA'
    try:
        Vacancy = table.find(string = re.compile('Vacancy')).find_next('td').text.strip()
    except:
        Vacancy = 'NA'
    try:
        Taxes = table.find(string = re.compile('Taxes')).find_next('td').text.strip()
    except:
        Taxes = 'NA'
    try:
        OPEX = table.find(string = re.compile('Operating Expenses')).find_next('td').text.strip()
    except:
        OPEX = 'NA'
    try:
        NOI = table.find(string = re.compile('Net Operating Income')).find_next('td').text.strip()
    except:
        NOI = 'NA'
        
    df4 = df4.append({'address':address, 'link':link, 'size':size, 'price':price, 'cap rate':cap_rate, 'number of units':num_units,
                    'year built':year_built, 'class':Class, 'GRI':GRI, 'vacancy':Vacancy, 'taxes':Taxes, 'OPEX':OPEX, 'NOI':NOI, 'zip code':zip_code}, ignore_index = True)
    next_link = soup.find('div', class_='paging')
    next_property = next_link.find_all('a')[1].get('href')
    driver.get('https://www.loopnet.com/' + next_property)
    soup = BeautifulSoup(driver.page_source, 'lxml')

with pd.ExcelWriter('OutputData.xlsx', mode= 'a', engine='openpyxl', if_sheet_exists='replace') as writer:
    df1.to_excel(writer, sheet_name='Sheet_1')
    df2.to_excel(writer, sheet_name='Sheet_2')
    df3.to_excel(writer, sheet_name='Sheet_3')
    df4.to_excel(writer, sheet_name='Sheet_4')


